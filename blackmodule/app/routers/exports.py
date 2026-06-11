from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from sqlalchemy import func
from app.models import Alert, SanctionEntry, AuditLog, MatchingSetting
from app.services.audit_service import write_audit_log


router = APIRouter(
    prefix="/api/exports",
    tags=["Exports"]
)


@router.get("/alerts-excel")
def export_alerts_excel(
    statut: str | None = None,
    niveau_alerte: str | None = None,
    source_liste: str | None = None,
    client_reference: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    critical_only: int | None = None,
    db: Session = Depends(get_db)
):
    query = db.query(Alert)
    if critical_only == 1:
        query = query.filter(
            Alert.niveau_alerte.in_(["ALERTE_EXACTE", "ALERTE_PROBABLE"]),
            Alert.statut.in_(["GENEREE", "EN_COURS", "ESCALADEE", "CONFIRMEE"])
        )

    if statut:
        statut_value = statut.strip().upper()
        query = query.filter(Alert.statut == statut_value)

    if niveau_alerte:
        niveau_value = niveau_alerte.strip().upper()
        query = query.filter(Alert.niveau_alerte == niveau_value)

    if source_liste:
        source_value = source_liste.strip().upper()
        query = query.filter(Alert.source_liste == source_value)

    if client_reference:
        client_ref_value = client_reference.strip()
        query = query.filter(Alert.client_reference.ilike(f"%{client_ref_value}%"))

    if date_from:
        try:
            parsed_date_from = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(Alert.created_at >= parsed_date_from)
        except ValueError:
            pass

    if date_to:
        try:
            parsed_date_to = datetime.strptime(date_to, "%Y-%m-%d")
            parsed_date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            query = query.filter(Alert.created_at <= parsed_date_to)
        except ValueError:
            pass

    alerts = query.order_by(Alert.created_at.desc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alertes BLACKMODULE"

    headers = [
        "Référence client",
        "Nom client",
        "Prénom client",
        "Date naissance",
        "Source liste",
        "Score matching",
        "Type matching",
        "Niveau alerte",
        "Statut",
        "Action recommandée",
        "Traité par",
        "Commentaire traitement",
        "Date création",
        "Date traitement"
    ]

    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for alert in alerts:
        sheet.append([
            alert.client_reference or "",
            alert.client_nom or "",
            alert.client_prenom or "",
            alert.client_date_naissance.strftime("%Y-%m-%d") if alert.client_date_naissance else "",
            alert.source_liste or "",
            float(alert.matching_score) if alert.matching_score is not None else "",
            alert.matching_type or "",
            alert.niveau_alerte or "",
            alert.statut or "",
            alert.action_recommandee or "",
            alert.treated_by or "",
            alert.treatment_comment or "",
            alert.created_at.strftime("%Y-%m-%d %H:%M:%S") if alert.created_at else "",
            alert.treated_at.strftime("%Y-%m-%d %H:%M:%S") if alert.treated_at else ""
        ])

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 45)

    filter_description = (
        f"Filtres appliqués - "
        f"Statut: {statut or 'Tous'}, "
        f"Niveau: {niveau_alerte or 'Tous'}, "
        f"Source: {source_liste or 'Toutes'}, "
        f"Référence client: {client_reference or 'Toutes'}, "
        f"Date début: {date_from or '-'}, "
        f"Date fin: {date_to or '-'}."
    )

    write_audit_log(
        db=db,
        user_identifier="SYSTEM",
        action="EXPORT_ALERTS_EXCEL",
        entity_type="Alert",
        entity_id=None,
        description=(
            f"Export Excel des alertes généré. "
            f"Nombre d'alertes exportées : {len(alerts)}. "
            f"{filter_description}"
        ),
        ip_address=None
    )

    db.commit()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"blackmodule_alertes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@router.get("/sanctions-excel")
def export_sanctions_excel(
    db: Session = Depends(get_db)
):
    sanctions = db.query(SanctionEntry).order_by(
        SanctionEntry.created_at.desc()
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sanctions BLACKMODULE"

    headers = [
        "Source liste",
        "Type entité",
        "Nom",
        "Prénom",
        "Nom complet",
        "Alias",
        "Date naissance",
        "Nationalité",
        "Pays",
        "Numéro passeport",
        "Motif sanction",
        "Date inscription",
        "Date suppression",
        "Statut",
        "Date création"
    ]

    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for sanction in sanctions:
        aliases = ""

        if sanction.aliases:
            aliases = "; ".join([
                alias.alias for alias in sanction.aliases
                if alias.alias
            ])

        sheet.append([
            sanction.source_liste,
            sanction.type_entite,
            sanction.nom,
            sanction.prenom,
            sanction.nom_complet,
            aliases,
            sanction.date_naissance.strftime("%Y-%m-%d") if sanction.date_naissance else "",
            sanction.nationalite,
            sanction.pays,
            sanction.num_passeport,
            sanction.motif_sanction,
            sanction.date_inscription.strftime("%Y-%m-%d") if sanction.date_inscription else "",
            sanction.date_suppression.strftime("%Y-%m-%d") if sanction.date_suppression else "",
            sanction.statut,
            sanction.created_at.strftime("%Y-%m-%d %H:%M:%S") if sanction.created_at else ""
        ])

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 50)

    write_audit_log(
        db=db,
        user_identifier="SYSTEM",
        action="EXPORT_SANCTIONS_EXCEL",
        entity_type="SanctionEntry",
        entity_id=None,
        description=f"Export Excel des sanctions/PPE généré. Nombre d'entrées exportées : {len(sanctions)}.",
        ip_address=None
    )

    db.commit()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"blackmodule_sanctions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@router.get("/audit-logs-excel")
def export_audit_logs_excel(
    action: str | None = None,
    user_identifier: str | None = None,
    entity_type: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    db: Session = Depends(get_db)
):
    query = db.query(AuditLog)

    if action:
        action_value = action.strip().upper()
        query = query.filter(AuditLog.action.ilike(f"%{action_value}%"))

    if user_identifier:
        user_value = user_identifier.strip()
        query = query.filter(AuditLog.user_identifier.ilike(f"%{user_value}%"))

    if entity_type:
        entity_value = entity_type.strip()
        query = query.filter(AuditLog.entity_type.ilike(f"%{entity_value}%"))

    if date_from:
        try:
            parsed_date_from = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(AuditLog.created_at >= parsed_date_from)
        except ValueError:
            pass

    if date_to:
        try:
            parsed_date_to = datetime.strptime(date_to, "%Y-%m-%d")
            parsed_date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            query = query.filter(AuditLog.created_at <= parsed_date_to)
        except ValueError:
            pass

    logs = query.order_by(
        AuditLog.created_at.desc()
    ).limit(5000).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Audit Logs BLACKMODULE"

    headers = [
        "Date",
        "Utilisateur",
        "Action",
        "Type entité",
        "ID entité",
        "Description",
        "Adresse IP"
    ]

    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for log in logs:
        sheet.append([
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
            log.user_identifier or "",
            log.action or "",
            log.entity_type or "",
            log.entity_id or "",
            log.description or "",
            log.ip_address or ""
        ])

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 70)

    filter_description = (
        f"Filtres appliqués - "
        f"Action: {action or 'Toutes'}, "
        f"Utilisateur: {user_identifier or 'Tous'}, "
        f"Entité: {entity_type or 'Toutes'}, "
        f"Date début: {date_from or '-'}, "
        f"Date fin: {date_to or '-'}."
    )

    write_audit_log(
        db=db,
        user_identifier="SYSTEM",
        action="EXPORT_AUDIT_LOGS_EXCEL",
        entity_type="AuditLog",
        entity_id=None,
        description=(
            f"Export Excel du journal d'audit généré. "
            f"Nombre de lignes exportées : {len(logs)}. "
            f"{filter_description}"
        ),
        ip_address=None
    )

    db.commit()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"blackmodule_audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@router.get("/data-quality-excel")
def export_data_quality_excel(
    db: Session = Depends(get_db)
):
    total_sanctions = db.query(SanctionEntry).count()

    missing_full_name = db.query(SanctionEntry).filter(
        (SanctionEntry.nom_complet == None) |
        (SanctionEntry.nom_complet == "")
    ).count()

    missing_source = db.query(SanctionEntry).filter(
        (SanctionEntry.source_liste == None) |
        (SanctionEntry.source_liste == "")
    ).count()

    missing_status = db.query(SanctionEntry).filter(
        (SanctionEntry.statut == None) |
        (SanctionEntry.statut == "")
    ).count()

    missing_hash = db.query(SanctionEntry).filter(
        (SanctionEntry.hash_signature == None) |
        (SanctionEntry.hash_signature == "")
    ).count()

    short_names = db.query(SanctionEntry).filter(
        SanctionEntry.nom_complet != None,
        SanctionEntry.nom_complet != "",
        func.length(SanctionEntry.nom_complet) < 3
    ).count()

    duplicate_hash_rows = db.query(
        SanctionEntry.hash_signature,
        func.count(SanctionEntry.id).label("count")
    ).filter(
        SanctionEntry.hash_signature != None,
        SanctionEntry.hash_signature != ""
    ).group_by(
        SanctionEntry.hash_signature
    ).having(
        func.count(SanctionEntry.id) > 1
    ).all()

    source_stats = db.query(
        SanctionEntry.source_liste,
        func.count(SanctionEntry.id).label("count")
    ).group_by(
        SanctionEntry.source_liste
    ).order_by(
        func.count(SanctionEntry.id).desc()
    ).all()

    problem_entries = db.query(SanctionEntry).filter(
        (
            (SanctionEntry.nom_complet == None) |
            (SanctionEntry.nom_complet == "") |
            (SanctionEntry.source_liste == None) |
            (SanctionEntry.source_liste == "") |
            (SanctionEntry.statut == None) |
            (SanctionEntry.statut == "") |
            (SanctionEntry.hash_signature == None) |
            (SanctionEntry.hash_signature == "")
        )
    ).order_by(
        SanctionEntry.created_at.desc()
    ).limit(5000).all()

    quality_score = 100

    if total_sanctions > 0:
        anomaly_total = (
            missing_full_name +
            missing_source +
            missing_status +
            missing_hash +
            short_names +
            len(duplicate_hash_rows)
        )

        quality_score = max(
            0,
            round(100 - ((anomaly_total / total_sanctions) * 100), 2)
        )

    workbook = Workbook()

    # Feuille 1 : Synthèse
    sheet = workbook.active
    sheet.title = "Synthese"

    headers = ["Indicateur", "Valeur"]
    sheet.append(headers)

    rows = [
        ["Total sanctions", total_sanctions],
        ["Score qualité", f"{quality_score}%"],
        ["Sans nom complet", missing_full_name],
        ["Sans source", missing_source],
        ["Sans statut", missing_status],
        ["Sans hash signature", missing_hash],
        ["Noms trop courts", short_names],
        ["Doublons hash", len(duplicate_hash_rows)]
    ]

    for row in rows:
        sheet.append(row)

    # Feuille 2 : Répartition par source
    source_sheet = workbook.create_sheet("Repartition par source")
    source_sheet.append(["Source", "Nombre"])

    for source, count in source_stats:
        source_sheet.append([
            source or "SOURCE VIDE",
            count
        ])

    # Feuille 3 : Entrées problématiques
    problem_sheet = workbook.create_sheet("Entrees problematiques")
    problem_sheet.append([
        "Source",
        "Nom complet",
        "Nom",
        "Prenom",
        "Statut",
        "Hash signature",
        "Date creation"
    ])

    for item in problem_entries:
        problem_sheet.append([
            item.source_liste or "",
            item.nom_complet or "",
            item.nom or "",
            item.prenom or "",
            item.statut or "",
            item.hash_signature or "",
            item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else ""
        ])

    # Feuille 4 : Doublons hash
    duplicate_sheet = workbook.create_sheet("Doublons hash")
    duplicate_sheet.append(["Hash signature", "Nombre"])

    for hash_signature, count in duplicate_hash_rows:
        duplicate_sheet.append([
            hash_signature,
            count
        ])

    # Style headers
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in workbook.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 70)

    write_audit_log(
        db=db,
        user_identifier="SYSTEM",
        action="EXPORT_DATA_QUALITY_EXCEL",
        entity_type="DataQuality",
        entity_id=None,
        description=(
            f"Export Excel du contrôle qualité généré. "
            f"Total sanctions : {total_sanctions}. "
            f"Score qualité : {quality_score}%."
        ),
        ip_address=None
    )

    db.commit()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"blackmodule_data_quality_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@router.get("/matching-settings-excel")
def export_matching_settings_excel(
    db: Session = Depends(get_db)
):
    settings = db.query(MatchingSetting).first()

    history_logs = db.query(AuditLog).filter(
        AuditLog.action == "UPDATE_MATCHING_SETTINGS"
    ).order_by(
        AuditLog.created_at.desc()
    ).limit(50).all()

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Parametres matching"

    sheet.append(["Paramètre", "Valeur"])

    if settings:
        rows = [
            ["Seuil Alerte Exacte", settings.exact_threshold],
            ["Seuil Alerte Probable", settings.probable_threshold],
            ["Seuil Alerte Possible", settings.possible_threshold],
            ["Dernière modification par", settings.updated_by or ""],
            [
                "Date dernière modification",
                settings.updated_at.strftime("%Y-%m-%d %H:%M:%S") if settings.updated_at else ""
            ],
        ]
    else:
        rows = [
            ["Seuil Alerte Exacte", 90],
            ["Seuil Alerte Probable", 75],
            ["Seuil Alerte Possible", 60],
            ["Dernière modification par", "SYSTEM"],
            ["Date dernière modification", ""],
        ]

    for row in rows:
        sheet.append(row)

    history_sheet = workbook.create_sheet("Historique")
    history_sheet.append([
        "Date",
        "Utilisateur",
        "Action",
        "Description",
        "Adresse IP"
    ])

    for log in history_logs:
        history_sheet.append([
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
            log.user_identifier or "",
            log.action or "",
            log.description or "",
            log.ip_address or "",
        ])

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in workbook.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 80)

    write_audit_log(
        db=db,
        user_identifier="SYSTEM",
        action="EXPORT_MATCHING_SETTINGS_EXCEL",
        entity_type="MatchingSetting",
        entity_id=str(settings.id) if settings else None,
        description="Export Excel des paramètres de matching généré.",
        ip_address=None
    )

    db.commit()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"blackmodule_matching_settings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )