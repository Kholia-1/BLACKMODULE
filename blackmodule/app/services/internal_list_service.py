"""Internal-list domain service (LOT 2C).

Internal records deliberately reuse ``SanctionEntry`` so the hardened matching
pipeline sees only records explicitly approved as active.  Approval requests
remain the single four-eyes workflow.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session

from app.models import InternalListHistory, SanctionAlias, SanctionEntry
from app.services.audit_service import write_audit_log

INTERNAL_CATEGORIES = {
    "ANIF": "ANIF",
    "JUDICIAIRE": "Liste judiciaire",
    "PPE_INTERNE": "Personnes politiquement exposées",
    "SURVEILLANCE_AFB": "Surveillance interne AFB",
}

DRAFT = "BROUILLON"
PENDING = "EN_ATTENTE_VALIDATION"
ACTIVE = "ACTIF"  # kept consistent with the existing matching SQL predicate
SUSPENDED = "SUSPENDUE"
DELISTED = "RADIEE"
LIFECYCLE_STATUSES = {DRAFT, PENDING, ACTIVE, SUSPENDED, DELISTED}
OP_INTERNAL_LIST_CHANGE = "INTERNAL_LIST_CHANGE"

EDITABLE_FIELDS = {
    "nom", "prenom", "nom_complet", "date_naissance", "lieu_naissance",
    "nationalite", "pays", "type_entite", "document_type", "document_number",
    "num_passeport", "autres_documents", "motif_sanction", "source_reference",
    "risk_level", "date_inscription", "date_suppression", "compliance_comment",
    "ppe_type", "ppe_function", "ppe_institution", "ppe_country",
    "ppe_function_start_date", "ppe_function_end_date", "ppe_status", "ppe_relationship",
    "source_record_id",
}

SENSITIVE_FIELDS = {
    "risk_level", "document_type", "document_number", "num_passeport",
    "autres_documents", "motif_sanction", "source_reference",
    "compliance_comment", "ppe_relationship",
}


class DuplicateInternalEntryError(ValueError):
    def __init__(self, duplicates: list[dict]):
        self.duplicates = duplicates
        super().__init__("Doublon déterministe potentiel détecté; création bloquée.")


def category_label(category: str) -> str:
    return INTERNAL_CATEGORIES.get((category or "").upper(), category or "Inconnue")


def _safe_json(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    raise TypeError(f"Unsupported value: {type(value)!r}")


def _normalise_values(values: dict) -> dict:
    result = {}
    for key, value in (values or {}).items():
        if key not in EDITABLE_FIELDS or value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
            if not value or value.casefold() in {"none", "null"}:
                continue
        if key in {"nom", "prenom", "nom_complet", "nationalite", "pays", "num_passeport"}:
            value = str(value).upper()
        if key.endswith("_date") or key in {"date_naissance", "date_inscription", "date_suppression"}:
            if isinstance(value, str):
                value = date.fromisoformat(value)
        result[key] = value
    return result


def _normalise_values_for_category(values: dict, category: str) -> dict:
    clean = _normalise_values(values)
    if category != "PPE_INTERNE":
        clean = {key: value for key, value in clean.items() if not key.startswith("ppe_")}
    return clean


def _normalise_update_values(values: dict, category: str) -> dict:
    """Normalise raw form values while distinguishing omission from clearing."""
    clean = _normalise_values_for_category(values, category)
    for key, value in (values or {}).items():
        if key not in EDITABLE_FIELDS or (category != "PPE_INTERNE" and key.startswith("ppe_")):
            continue
        if isinstance(value, str) and not value.strip():
            clean[key] = None
    return clean


def _normalise_stored_update_values(values: dict, category: str) -> dict:
    """Validate an already-computed diff, preserving explicit field clearing."""
    clean = {}
    for key, value in (values or {}).items():
        if key not in EDITABLE_FIELDS or (category != "PPE_INTERNE" and key.startswith("ppe_")):
            continue
        if value is None:
            clean[key] = None
            continue
        normalised = _normalise_values({key: value})
        if key in normalised:
            clean[key] = normalised[key]
    return clean


def _normalise_aliases(aliases: list[str]) -> list[str]:
    result = []
    for alias in aliases:
        clean = str(alias).strip().upper()
        if clean and clean.casefold() not in {"none", "null"} and clean not in result:
            result.append(clean)
    return result


def entry_snapshot(entry: SanctionEntry) -> dict:
    return {
        field: getattr(entry, field) for field in EDITABLE_FIELDS
    } | {
        "source_liste": entry.source_liste,
        "internal_status": entry.internal_status,
        "aliases": [alias.alias for alias in entry.aliases],
    }


def serialize_internal_entry(entry: SanctionEntry, *, include_sensitive: bool) -> dict:
    """Build an explicit API payload; never serialize the ORM object directly."""
    payload = {
        "id": str(entry.id), "category": entry.source_liste, "source_liste": entry.source_liste,
        "type_entite": entry.type_entite, "nom": entry.nom, "prenom": entry.prenom,
        "nom_complet": entry.nom_complet, "aliases": [alias.alias for alias in entry.aliases],
        "date_naissance": entry.date_naissance, "lieu_naissance": entry.lieu_naissance,
        "nationalite": entry.nationalite, "pays": entry.pays,
        "date_inscription": entry.date_inscription, "date_suppression": entry.date_suppression,
        "status": entry.internal_status, "internal_status": entry.internal_status,
        "created_at": entry.created_at, "updated_at": entry.updated_at,
        "ppe_type": entry.ppe_type, "ppe_function": entry.ppe_function,
        "ppe_institution": entry.ppe_institution, "ppe_country": entry.ppe_country,
        "ppe_function_start_date": entry.ppe_function_start_date,
        "ppe_function_end_date": entry.ppe_function_end_date, "ppe_status": entry.ppe_status,
    }
    if include_sensitive:
        payload.update({field: getattr(entry, field) for field in SENSITIVE_FIELDS})
        payload["source_record_id"] = entry.source_record_id
        payload.update({
            "created_by": entry.created_by, "updated_by": entry.updated_by,
            "submitted_by": entry.submitted_by, "submitted_at": entry.submitted_at,
            "validated_by": entry.validated_by, "validated_at": entry.validated_at,
        })
        payload["history"] = [
            {
                "action": item.action, "performed_by": item.performed_by,
                "old_values": item.old_values, "new_values": item.new_values,
                "comment": item.comment, "created_at": item.created_at,
            }
            for item in getattr(entry, "internal_history", [])
        ]
    return payload


def find_internal_duplicates(
    db: Session, *, category: str, values: dict, exclude_entry_id=None,
) -> list[dict]:
    """Find deterministic duplicates only; fuzzy/name-only merging is forbidden."""
    clean = _normalise_values(values)
    conditions = []
    source_record_id = clean.get("source_record_id")
    source_reference = clean.get("source_reference")
    document_number = clean.get("document_number")
    passport = clean.get("num_passeport")
    birth_date = clean.get("date_naissance")
    name = clean.get("nom")
    if source_record_id:
        conditions.append(and_(SanctionEntry.source_liste == category, SanctionEntry.source_record_id == source_record_id))
    elif source_reference:
        conditions.append(and_(SanctionEntry.source_liste == category, func.upper(SanctionEntry.source_reference) == str(source_reference).strip().upper()))
    if document_number:
        conditions.append(func.upper(SanctionEntry.document_number) == str(document_number).strip().upper())
    if passport:
        conditions.append(func.upper(SanctionEntry.num_passeport) == str(passport).strip().upper())
    if name and birth_date:
        conditions.append(and_(func.upper(SanctionEntry.nom) == name, SanctionEntry.date_naissance == birth_date))
    if not conditions:
        return []
    query = db.query(SanctionEntry).filter(
        SanctionEntry.is_internal_list.is_(True), or_(*conditions)
    )
    if exclude_entry_id:
        query = query.filter(SanctionEntry.id != uuid.UUID(str(exclude_entry_id)))
    return [
        {"id": str(item.id), "category": item.source_liste, "status": item.internal_status}
        for item in query.limit(20).all()
    ]


def _history(db: Session, entry: SanctionEntry, action: str, actor: str | None,
             old_values: dict | None = None, new_values: dict | None = None,
             comment: str | None = None, approval_id=None) -> None:
    db.add(InternalListHistory(
        sanction_entry_id=entry.id, action=action, performed_by=actor,
        old_values=json.dumps(old_values, ensure_ascii=False, default=_safe_json) if old_values else None,
        new_values=json.dumps(new_values, ensure_ascii=False, default=_safe_json) if new_values else None,
        comment=comment, approval_request_id=approval_id,
    ))


def _audit(db: Session, actor: str | None, action: str, entry: SanctionEntry) -> None:
    # Never log identity attributes / document numbers / motives.
    write_audit_log(db, actor, action, "InternalSanctionEntry", str(entry.id),
                    f"Opération {action} sur une fiche interne {category_label(entry.source_liste)}.", None)


def create_internal_entry(db: Session, *, category: str, values: dict, aliases: list[str], actor: str) -> SanctionEntry:
    category = (category or "").upper()
    if category not in INTERNAL_CATEGORIES:
        raise ValueError("Catégorie de liste interne invalide.")
    clean = _normalise_values_for_category(values, category)
    if not clean.get("nom"):
        raise ValueError("Le nom est obligatoire.")
    duplicates = find_internal_duplicates(db, category=category, values=clean)
    if duplicates:
        raise DuplicateInternalEntryError(duplicates)
    clean.setdefault("type_entite", "PERSONNE_PHYSIQUE")
    clean.setdefault("nom_complet", " ".join(part for part in [clean.get("prenom"), clean["nom"]] if part))
    entry = SanctionEntry(
        source_liste=category, is_internal_list=True, internal_status=DRAFT,
        statut=DRAFT, created_by=actor, updated_by=actor, **clean,
    )
    db.add(entry)
    db.flush()
    _replace_aliases(entry, aliases)
    _history(db, entry, "CREATION", actor, new_values=entry_snapshot(entry))
    _audit(db, actor, "INTERNAL_LIST_CREATED", entry)
    return entry


def _replace_aliases(entry: SanctionEntry, aliases: list[str]) -> None:
    entry.aliases[:] = []
    for clean in _normalise_aliases(aliases or []):
        entry.aliases.append(SanctionAlias(alias=clean))


def create_change_request(db: Session, *, entry: SanctionEntry, actor: dict, action: str,
                          values: dict | None = None, aliases: list[str] | None = None,
                          comment: str | None = None, ip_address: str | None = None):
    """Create a four-eyes request; no active value is mutated before approval."""
    from app.services.approval_service import create_approval_request

    if not entry.is_internal_list:
        raise ValueError("La fiche cible n'est pas une fiche interne.")
    if action not in {"ACTIVATE", "UPDATE", "SUSPEND", "REACTIVATE", "RADIATE"}:
        raise ValueError("Transition interne invalide.")
    from app.models import ApprovalRequest
    existing_pending = db.query(ApprovalRequest).filter(
        ApprovalRequest.operation_type == OP_INTERNAL_LIST_CHANGE,
        ApprovalRequest.target_entity_id == str(entry.id),
        ApprovalRequest.status == PENDING,
    ).first()
    if existing_pending:
        raise ValueError("Une demande de validation est déjà en attente pour cette fiche.")
    old = entry_snapshot(entry)
    payload_values = (
        _normalise_stored_update_values(values or {}, entry.source_liste)
        if action == "UPDATE" else _normalise_values_for_category(values or {}, entry.source_liste)
    )
    payload = {"action": action, "values": payload_values}
    if aliases is not None:
        payload["aliases"] = _normalise_aliases(aliases)
    approval = create_approval_request(
        db, operation_type=OP_INTERNAL_LIST_CHANGE, initiator=actor,
        target_entity_type="InternalSanctionEntry", target_entity_id=str(entry.id),
        old_values=old, new_values=payload, comment=comment, ip_address=ip_address,
    )
    _history(db, entry, "SOUMISSION", actor.get("username"), old_values=old,
             new_values=payload, comment=comment, approval_id=approval.id)
    _audit(db, actor.get("username"), "INTERNAL_LIST_SUBMITTED", entry)
    return approval


def submit_internal_entry(db: Session, *, entry: SanctionEntry, actor: dict,
                          comment: str | None, ip_address: str | None):
    if entry.internal_status != DRAFT:
        raise ValueError("Seule une fiche brouillon peut être soumise.")
    entry.internal_status = PENDING
    entry.statut = PENDING
    entry.submitted_by = actor.get("username")
    entry.submitted_at = datetime.utcnow()
    return create_change_request(db, entry=entry, actor=actor, action="ACTIVATE",
                                 comment=comment, ip_address=ip_address)


def request_entry_change(db: Session, *, entry: SanctionEntry, actor: dict, action: str,
                         values: dict | None, aliases: list[str] | None,
                         comment: str | None, ip_address: str | None):
    if entry.internal_status == PENDING:
        raise ValueError("Une demande de validation est déjà en attente pour cette fiche.")
    valid_origins = {
        "UPDATE": {DRAFT, ACTIVE},
        "SUSPEND": {ACTIVE},
        "REACTIVATE": {SUSPENDED},
        "RADIATE": {ACTIVE, SUSPENDED},
    }
    if action not in valid_origins or entry.internal_status not in valid_origins[action]:
        raise ValueError("Transition incompatible avec le statut actuel de la fiche.")
    changed_values = {}
    changed_aliases = None
    if action == "UPDATE":
        normalised = _normalise_update_values(values or {}, entry.source_liste)
        changed_values = {
            key: value for key, value in normalised.items() if getattr(entry, key) != value
        }
        if aliases is not None:
            normalised_aliases = _normalise_aliases(aliases)
            current_aliases = [alias.alias for alias in entry.aliases]
            if normalised_aliases != current_aliases:
                changed_aliases = normalised_aliases
        if not changed_values and changed_aliases is None:
            raise ValueError("Aucune modification métier détectée.")
        proposed = entry_snapshot(entry)
        proposed.update(changed_values)
        duplicates = find_internal_duplicates(
            db, category=entry.source_liste, values=proposed, exclude_entry_id=entry.id,
        )
        if duplicates:
            raise DuplicateInternalEntryError(duplicates)
    if action == "UPDATE" and entry.internal_status == DRAFT:
        old = entry_snapshot(entry)
        for key, value in changed_values.items():
            setattr(entry, key, value)
        if changed_aliases is not None:
            _replace_aliases(entry, changed_aliases)
        entry.updated_by = actor.get("username")
        _history(db, entry, "MODIFICATION_BROUILLON", actor.get("username"), old, entry_snapshot(entry), comment)
        _audit(db, actor.get("username"), "INTERNAL_LIST_UPDATED", entry)
        return None
    return create_change_request(db, entry=entry, actor=actor, action=action,
                                 values=changed_values if action == "UPDATE" else values,
                                 aliases=changed_aliases if action == "UPDATE" else aliases,
                                 comment=comment, ip_address=ip_address)


def apply_approved_internal_change(db: Session, *, entry_id: str, values: dict,
                                   reviewer_username: str, approval_id) -> None:
    entry = db.query(SanctionEntry).filter(
        SanctionEntry.id == uuid.UUID(str(entry_id))
    ).with_for_update().first()
    if not entry or not entry.is_internal_list:
        raise ValueError("Fiche interne cible introuvable.")
    old = entry_snapshot(entry)
    action = values.get("action")
    if action == "ACTIVATE":
        entry.internal_status = ACTIVE
        entry.statut = ACTIVE
    elif action == "UPDATE":
        for key, value in _normalise_stored_update_values(
            values.get("values") or {}, entry.source_liste
        ).items():
            setattr(entry, key, value)
        if "aliases" in values:
            _replace_aliases(entry, values["aliases"])
    elif action == "SUSPEND":
        entry.internal_status = SUSPENDED
        entry.statut = SUSPENDED
    elif action == "REACTIVATE":
        entry.internal_status = ACTIVE
        entry.statut = ACTIVE
    elif action == "RADIATE":
        entry.internal_status = DELISTED
        entry.statut = DELISTED
        entry.date_suppression = date.today()
    else:
        raise ValueError("Action interne inconnue.")
    entry.updated_by = reviewer_username
    entry.validated_by = reviewer_username
    entry.validated_at = datetime.utcnow()
    _history(db, entry, f"VALIDATION_{action}", reviewer_username, old, entry_snapshot(entry), approval_id=approval_id)
    _audit(db, reviewer_username, f"INTERNAL_LIST_{action}_APPROVED", entry)


def apply_rejected_internal_change(db: Session, approval, reviewer_username: str) -> None:
    if approval.operation_type != OP_INTERNAL_LIST_CHANGE:
        return
    entry = db.query(SanctionEntry).filter(
        SanctionEntry.id == uuid.UUID(str(approval.target_entity_id))
    ).first()
    if not entry:
        return
    values = json.loads(approval.new_values or "{}")
    if values.get("action") == "ACTIVATE" and entry.internal_status == PENDING:
        entry.internal_status = DRAFT
        entry.statut = DRAFT
    _history(db, entry, "REJET", reviewer_username, comment=approval.reviewer_comment, approval_id=approval.id)
    _audit(db, reviewer_username, "INTERNAL_LIST_REJECTED", entry)


IMPORT_COLUMNS = {"nom"}
IMPORT_OPTIONAL_COLUMNS = EDITABLE_FIELDS | {"alias", "aliases"}


def parse_internal_import(file_content: bytes, filename: str) -> tuple[list[dict], list[dict]]:
    """Return accepted and rejected rows without persisting anything."""
    filename = (filename or "").lower()
    if filename.endswith(".csv"):
        rows = list(csv.DictReader(io.StringIO(file_content.decode("utf-8-sig"))))
    elif filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
        rows = [dict(zip(headers, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]
    else:
        raise ValueError("Seuls les fichiers CSV et XLSX sont acceptés.")
    if not rows:
        return [], []
    headers = {str(key).strip().lower() for key in rows[0] if key}
    missing = IMPORT_COLUMNS - headers
    unknown = headers - {column.lower() for column in IMPORT_OPTIONAL_COLUMNS}
    if missing or unknown:
        messages = []
        if missing:
            messages.append("colonne obligatoire absente : " + ", ".join(sorted(missing)))
        if unknown:
            messages.append("colonne non reconnue : " + ", ".join(sorted(unknown)))
        raise ValueError("; ".join(messages))
    accepted, rejected = [], []
    for index, row in enumerate(rows, start=2):
        normalized = {(key or "").strip().lower(): value for key, value in row.items()}
        if not str(normalized.get("nom") or "").strip():
            rejected.append({"row": index, "error": "nom obligatoire", "values": normalized})
            continue
        try:
            _normalise_values(normalized)
        except (TypeError, ValueError):
            rejected.append({"row": index, "error": "date ou valeur invalide", "values": normalized})
            continue
        # Preserve the source line for the preview UI.  The metadata is ignored
        # by the existing value normalisation before any record is persisted.
        normalized["__import_row__"] = index
        accepted.append(normalized)
    return accepted, rejected


def preview_internal_import(db: Session, *, category: str, file_content: bytes, filename: str):
    accepted, rejected = parse_internal_import(file_content, filename)
    clean_accepted, duplicates = [], []
    seen_keys: dict[tuple, int] = {}
    for index, row in enumerate(accepted, start=2):
        row_number = int(row.get("__import_row__", index))
        normalized = _normalise_values(row)
        keys = []
        if normalized.get("source_record_id"):
            keys.append(("source_record_id", category, normalized["source_record_id"]))
        elif normalized.get("source_reference"):
            keys.append(("source_reference", category, str(normalized["source_reference"]).upper()))
        if normalized.get("document_number"):
            keys.append(("document_number", str(normalized["document_number"]).upper()))
        if normalized.get("num_passeport"):
            keys.append(("passport", normalized["num_passeport"]))
        if normalized.get("nom") and normalized.get("date_naissance"):
            keys.append(("identity", normalized["nom"], normalized["date_naissance"].isoformat()))
        repeated = next((seen_keys[key] for key in keys if key in seen_keys), None)
        database_matches = find_internal_duplicates(db, category=category, values=normalized)
        if repeated is not None or database_matches:
            duplicates.append({
                "row": row_number,
                "reason": f"doublon dans le fichier (ligne {repeated})" if repeated is not None else "doublon avec une fiche existante",
                "matches": database_matches,
            })
            continue
        for key in keys:
            seen_keys[key] = row_number
        clean_accepted.append(row)
    return clean_accepted, rejected, duplicates
