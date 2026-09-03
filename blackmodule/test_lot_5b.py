"""LOT 5B regression tests for management reporting."""

import unittest
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import FastAPI, Request
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.middleware.sessions import SessionMiddleware

from app.database import Base, get_db
from app.models import Alert, AlertDecisionHistory, AuditLog, User
from app.routers import web
from app.services.alert_decision_service import APPROVED_DECISION
from app.services.authorization_service import (
    ROLE_ANALYSTE_CONFORMITE,
    ROLE_AUDITEUR,
    ROLE_SUPERVISEUR_CONFORMITE,
)
from app.services.reporting_service import (
    build_management_report,
    previous_reporting_filters,
    resolve_reporting_filters,
)


def _test_app(db) -> FastAPI:
    app = FastAPI()
    app.add_middleware(SessionMiddleware, secret_key="lot-5b-local-test")

    @app.get("/_test/login/{role}")
    def login(request: Request, role: str):
        user = db.query(User).filter(User.role == role).first()
        request.session["user"] = {
            "id": str(user.id), "username": user.username,
            "full_name": user.full_name, "role": user.role,
        }
        return {"status": "ok"}

    app.dependency_overrides[get_db] = lambda: db
    app.include_router(web.router)
    return app


class Lot5BManagementReportingTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False}, poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.db = sessionmaker(bind=self.engine, expire_on_commit=False)()
        self.now = datetime(2026, 9, 3, 12, 0)
        self.analyst1 = self.user("analyste-management-1", ROLE_ANALYSTE_CONFORMITE, "Alice Analyste")
        self.analyst2 = self.user("analyste-management-2", ROLE_ANALYSTE_CONFORMITE, "Bruno Analyste")
        self.supervisor = self.user("superviseur-management", ROLE_SUPERVISEUR_CONFORMITE, "Sophie Superviseur")
        self.auditor = self.user("auditeur-management", ROLE_AUDITEUR, "Arthur Auditeur")
        self.alerts = self.seed_alerts()
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.clear_compiled_cache()
        self.engine.dispose()

    def user(self, username, role, full_name):
        user = User(
            username=username, full_name=full_name,
            email=f"{username}@example.test", password_hash="local-test",
            role=role, statut="ACTIF",
        )
        self.db.add(user)
        self.db.flush()
        return user

    def alert(
        self, reference, *, created_at, status, level, source,
        matching_type, analyst=None, treated_at=None,
    ):
        item = Alert(
            client_reference=reference,
            client_nom="CLIENT MANAGEMENT CONFIDENTIEL",
            source_liste=source,
            matching_score=92,
            matching_type=matching_type,
            niveau_alerte=level,
            statut=status,
            created_at=created_at,
            assigned_to_user_id=analyst.id if analyst else None,
            assigned_to=analyst.username if analyst else None,
            assigned_at=created_at if analyst else None,
            treated_at=treated_at,
            treated_by=analyst.username if analyst and treated_at else None,
            treatment_comment="COMMENTAIRE MANAGEMENT CONFIDENTIEL" if treated_at else None,
        )
        self.db.add(item)
        self.db.flush()
        return item

    def decision(self, alert, status, applied_at):
        self.db.add(AlertDecisionHistory(
            alert_id=alert.id,
            old_status="EN_COURS",
            requested_status=status,
            decision_status=APPROVED_DECISION,
            initiated_by=alert.assigned_to or self.analyst1.username,
            initiated_at=applied_at - timedelta(hours=1),
            reason="MOTIF MANAGEMENT CONFIDENTIEL",
            reviewed_by=self.supervisor.username,
            reviewed_at=applied_at,
            applied_at=applied_at,
        ))

    def seed_alerts(self):
        current_critical = self.alert(
            "CURRENT-CRITICAL", created_at=self.now - timedelta(hours=2),
            status="GENEREE", level="ALERTE_EXACTE", source="OFAC_SDN",
            matching_type="EXACT_NAME",
        )
        current_out_sla = self.alert(
            "CURRENT-OUT-SLA", created_at=self.now - timedelta(hours=30),
            status="EN_COURS", level="ALERTE_PROBABLE", source="ONU",
            matching_type="FUZZY_NAME", analyst=self.analyst1,
        )
        current_closed = self.alert(
            "CURRENT-CLOSED", created_at=self.now - timedelta(hours=48),
            status="CLOTUREE", level="ALERTE_PROBABLE", source="OFAC_SDN",
            matching_type="NAME_ABBREVIATION", analyst=self.analyst1,
            treated_at=self.now - timedelta(hours=2),
        )
        self.decision(current_closed, "CLOTUREE", self.now - timedelta(hours=2))
        current_false_positive = self.alert(
            "CURRENT-FP", created_at=self.now - timedelta(hours=96),
            status="FAUX_POSITIF", level="ALERTE_POSSIBLE", source="ONU",
            matching_type="FUZZY_NAME", analyst=self.analyst2,
            treated_at=self.now - timedelta(hours=24),
        )
        self.decision(current_false_positive, "FAUX_POSITIF", self.now - timedelta(hours=24))
        current_confirmed = self.alert(
            "CURRENT-CONFIRMED", created_at=self.now - timedelta(hours=24),
            status="CONFIRMEE", level="ALERTE_POSSIBLE", source="OFAC_SDN",
            matching_type="EXACT_DOCUMENT", analyst=self.analyst2,
            treated_at=self.now - timedelta(hours=1),
        )
        self.decision(current_confirmed, "CONFIRMEE", self.now - timedelta(hours=1))

        previous_closed = self.alert(
            "PREVIOUS-CLOSED", created_at=self.now - timedelta(days=12),
            status="CLOTUREE", level="ALERTE_PROBABLE", source="OFAC_SDN",
            matching_type="EXACT_NAME", analyst=self.analyst1,
            treated_at=self.now - timedelta(days=9),
        )
        self.decision(previous_closed, "CLOTUREE", self.now - timedelta(days=9))
        previous_confirmed = self.alert(
            "PREVIOUS-CONFIRMED", created_at=self.now - timedelta(days=11),
            status="CONFIRMEE", level="ALERTE_PROBABLE", source="ONU",
            matching_type="FUZZY_NAME", analyst=self.analyst2,
            treated_at=self.now - timedelta(days=8),
        )
        self.decision(previous_confirmed, "CONFIRMEE", self.now - timedelta(days=8))
        old_critical = self.alert(
            "OLD-CRITICAL", created_at=self.now - timedelta(days=24),
            status="GENEREE", level="ALERTE_EXACTE", source="OFAC_SDN",
            matching_type="EXACT_NAME",
        )
        closed_during_current = self.alert(
            "CLOSED-DURING-CURRENT", created_at=self.now - timedelta(days=14),
            status="CLOTUREE", level="ALERTE_POSSIBLE", source="UE",
            matching_type="NORMALIZED_NAME", analyst=self.analyst1,
            treated_at=self.now - timedelta(days=2),
        )
        self.decision(closed_during_current, "CLOTUREE", self.now - timedelta(days=2))
        return [
            current_critical, current_out_sla, current_closed, current_false_positive,
            current_confirmed, previous_closed, previous_confirmed, old_critical,
            closed_during_current,
        ]

    def report(self, **kwargs):
        filters = resolve_reporting_filters(now=self.now, **kwargs)
        return build_management_report(self.db, filters, now=self.now)

    def test_01_management_kpis_and_rates_are_exact(self):
        report = self.report(period="7")
        current = report["current"]
        self.assertEqual(current["backlog_total"], 5)
        self.assertEqual(current["backlog_critical"], 2)
        self.assertEqual(current["backlog_out_sla"], 3)
        self.assertEqual(current["created"], 5)
        self.assertEqual(current["closed"], 2)
        self.assertEqual(current["closure_rate"], 40.0)
        self.assertEqual(current["false_positive_rate"], 50.0)
        self.assertEqual(current["confirmation_rate"], 50.0)
        self.assertEqual(current["sla_compliance_rate"], 40.0)

    def test_02_previous_period_and_variations_are_equivalent(self):
        report = self.report(period="7")
        previous = report["previous"]
        self.assertEqual((previous["backlog_total"], previous["backlog_critical"]), (3, 1))
        self.assertEqual(previous["backlog_out_sla"], 3)
        self.assertEqual((previous["created"], previous["closed"]), (2, 1))
        self.assertEqual(previous["closure_rate"], 50.0)
        self.assertEqual(previous["confirmation_rate"], 100.0)
        self.assertEqual(report["comparisons"]["backlog_total"]["absolute"], 2.0)
        self.assertEqual(report["comparisons"]["backlog_total"]["percent"], 66.7)
        filters = previous_reporting_filters(report["filters"])
        self.assertEqual(filters.end_at, report["filters"].start_at)
        self.assertEqual(filters.end_at - filters.start_at, report["filters"].end_at - report["filters"].start_at)

    def test_03_analysis_dimensions_and_backlog_age_are_correct(self):
        report = self.report(period="7")
        self.assertEqual(
            [(item["label"], item["count"]) for item in report["top_sources"]],
            [("OFAC_SDN", 3), ("ONU", 2)],
        )
        self.assertEqual(sum(item["count"] for item in report["current"]["age_buckets"]), 5)
        analysts = {item["username"]: item for item in report["analysts"]}
        self.assertEqual(analysts["analyste-management-1"]["closed"], 2)
        self.assertEqual(analysts["analyste-management-1"]["out_sla"], 1)
        self.assertNotIn("superviseur-management", analysts)
        self.assertTrue(report["attention_points"])

    def test_04_custom_and_source_filters_apply_to_both_periods(self):
        onu = self.report(period="7", source="ONU")
        self.assertEqual(onu["current"]["created"], 2)
        self.assertEqual(onu["current"]["backlog_total"], 2)
        self.assertEqual(onu["previous"]["created"], 1)
        custom = self.report(
            period="custom", date_from="2026-08-30", date_to="2026-09-02",
        )
        self.assertEqual((custom["filters"].end_at - custom["filters"].start_at).days, 4)
        self.assertEqual((custom["previous_filters"].end_at - custom["previous_filters"].start_at).days, 4)

    def test_05_rbac_page_and_navbar_are_backend_enforced(self):
        with TestClient(_test_app(self.db)) as client:
            for role in (ROLE_SUPERVISEUR_CONFORMITE, ROLE_AUDITEUR):
                client.get(f"/_test/login/{role}")
                response = client.get("/web/management-reporting?period=7")
                self.assertEqual(response.status_code, 200)
                self.assertIn("Synthèse management", response.text)
                self.assertIn('/web/management-reporting', response.text)
                self.assertNotIn("CLIENT MANAGEMENT CONFIDENTIEL", response.text)
            client.get(f"/_test/login/{ROLE_ANALYSTE_CONFORMITE}")
            self.assertEqual(client.get("/web/management-reporting?period=7").status_code, 403)
            self.assertEqual(client.get("/web/management-reporting/export.csv?period=7").status_code, 403)
            self.assertEqual(client.get("/web/management-reporting/export.xlsx?period=7").status_code, 403)

    def test_06_csv_and_xlsx_exports_are_filtered_and_non_sensitive(self):
        with TestClient(_test_app(self.db)) as client:
            client.get(f"/_test/login/{ROLE_SUPERVISEUR_CONFORMITE}")
            csv_response = client.get("/web/management-reporting/export.csv?period=7&source=ONU")
            self.assertEqual(csv_response.status_code, 200)
            self.assertIn("text/csv", csv_response.headers["content-type"])
            self.assertIn("Alertes créées;2", csv_response.text)
            xlsx_response = client.get("/web/management-reporting/export.xlsx?period=7&source=ONU")
            self.assertEqual(xlsx_response.status_code, 200)
            self.assertIn("spreadsheetml", xlsx_response.headers["content-type"])
            workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True, data_only=True)
            values = "\n".join(
                " | ".join("" if value is None else str(value) for value in row)
                for row in workbook.active.iter_rows(values_only=True)
            )
            for content in (csv_response.text, values):
                self.assertNotIn("CLIENT MANAGEMENT CONFIDENTIEL", content)
                self.assertNotIn("COMMENTAIRE MANAGEMENT CONFIDENTIEL", content)
                self.assertNotIn("MOTIF MANAGEMENT CONFIDENTIEL", content)
            actions = {
                row[0] for row in self.db.query(AuditLog.action)
                .filter(AuditLog.action.like("EXPORT_MANAGEMENT_REPORT_%")).all()
            }
            self.assertEqual(actions, {"EXPORT_MANAGEMENT_REPORT_CSV", "EXPORT_MANAGEMENT_REPORT_XLSX"})

    def test_07_management_query_count_is_constant_and_workflow_read_only(self):
        statuses_before = {str(alert.id): alert.statut for alert in self.alerts}
        history_before = self.db.query(AlertDecisionHistory).count()
        query_count = 0

        def count_query(*_args, **_kwargs):
            nonlocal query_count
            query_count += 1

        event.listen(self.engine, "before_cursor_execute", count_query)
        try:
            report = self.report(period="7")
        finally:
            event.remove(self.engine, "before_cursor_execute", count_query)
        self.assertEqual(query_count, 18)
        self.assertEqual({str(alert.id): alert.statut for alert in self.alerts}, statuses_before)
        self.assertEqual(self.db.query(AlertDecisionHistory).count(), history_before)
        serialized = str(report)
        self.assertNotIn("CLIENT MANAGEMENT CONFIDENTIEL", serialized)
        self.assertNotIn("COMMENTAIRE MANAGEMENT CONFIDENTIEL", serialized)
        self.assertNotIn("MOTIF MANAGEMENT CONFIDENTIEL", serialized)

    def test_08_null_status_remains_in_active_backlog(self):
        nullable = self.alert(
            "NULL-STATUS", created_at=self.now - timedelta(hours=1),
            status="GENEREE", level="ALERTE_POSSIBLE", source="OFAC_SDN",
            matching_type="FUZZY_NAME",
        )
        nullable.statut = None
        self.db.commit()
        self.assertEqual(self.report(period="7")["current"]["backlog_total"], 6)

    def test_09_reporting_never_materializes_alert_entities(self):
        loaded_alerts = 0

        def count_load(*_args, **_kwargs):
            nonlocal loaded_alerts
            loaded_alerts += 1

        self.db.expunge_all()
        event.listen(Alert, "load", count_load)
        try:
            self.report(period="7")
        finally:
            event.remove(Alert, "load", count_load)
        self.assertEqual(loaded_alerts, 0)

    def test_10_period_boundaries_are_half_open_and_non_overlapping(self):
        filters = resolve_reporting_filters(now=self.now, period="7")
        at_start = self.alert(
            "CREATED-AT-START", created_at=filters.start_at,
            status="GENEREE", level="ALERTE_POSSIBLE", source="UE",
            matching_type="FUZZY_NAME",
        )
        closed_at_start = self.alert(
            "CLOSED-AT-START", created_at=filters.start_at - timedelta(days=2),
            status="CLOTUREE", level="ALERTE_POSSIBLE", source="UE",
            matching_type="NORMALIZED_NAME", analyst=self.analyst1,
            treated_at=filters.start_at,
        )
        self.decision(closed_at_start, "CLOTUREE", filters.start_at)
        self.db.commit()
        report = build_management_report(self.db, filters, now=self.now)
        self.assertEqual(report["current"]["created"], 6)
        self.assertEqual(report["previous"]["created"], 3)
        self.assertEqual(report["current"]["closed"], 3)
        self.assertEqual(report["previous"]["closed"], 1)
        self.assertEqual(report["previous"]["backlog_total"], 4)
        self.assertIn(at_start, self.db)

    def test_11_legacy_treated_at_fallback_preserves_historical_backlog(self):
        late_terminal = self.alert(
            "LEGACY-LATE-CLOSE", created_at=self.now - timedelta(days=20),
            status="CLOTUREE", level="ALERTE_POSSIBLE", source="UE",
            matching_type="NORMALIZED_NAME", analyst=self.analyst1,
            treated_at=self.now - timedelta(days=2),
        )
        unknown_terminal = self.alert(
            "LEGACY-UNKNOWN-CLOSE", created_at=self.now - timedelta(days=20),
            status="CLOTUREE", level="ALERTE_POSSIBLE", source="UE",
            matching_type="NORMALIZED_NAME", analyst=self.analyst1,
        )
        self.db.commit()
        report = self.report(period="7")
        self.assertEqual(report["current"]["backlog_total"], 5)
        self.assertEqual(report["previous"]["backlog_total"], 4)
        self.assertIsNotNone(late_terminal.treated_at)
        self.assertIsNone(unknown_terminal.treated_at)

    def test_12_empty_period_zero_rates_and_unfiltered_xlsx_are_safe(self):
        report = self.report(
            period="custom", date_from="2025-01-01", date_to="2025-01-07",
        )
        self.assertEqual(report["current"]["created"], 0)
        self.assertEqual(report["current"]["closed"], 0)
        self.assertEqual(report["current"]["backlog_total"], 0)
        self.assertEqual(report["current"]["closure_rate"], 0.0)
        self.assertEqual(report["current"]["false_positive_rate"], 0.0)
        self.assertEqual(report["current"]["confirmation_rate"], 0.0)
        self.assertEqual(report["current"]["sla_compliance_rate"], 100.0)
        self.assertEqual(report["comparisons"]["created"]["percent"], 0.0)
        with TestClient(_test_app(self.db)) as client:
            client.get(f"/_test/login/{ROLE_AUDITEUR}")
            response = client.get(
                "/web/management-reporting/export.xlsx"
                "?period=custom&date_from=2025-01-01&date_to=2025-01-07"
            )
            self.assertEqual(response.status_code, 200)
            workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            self.assertGreater(len(rows), 10)
            self.assertIn(("Alertes créées", 0, 0, 0, 0), rows)
            broad_response = client.get("/web/management-reporting/export.xlsx?period=30")
            self.assertEqual(broad_response.status_code, 200)
            broad_workbook = load_workbook(
                BytesIO(broad_response.content), read_only=True, data_only=True,
            )
            broad_values = "\n".join(
                " | ".join("" if value is None else str(value) for value in row)
                for row in broad_workbook.active.iter_rows(values_only=True)
            )
            self.assertIn("Alertes créées", broad_values)
            self.assertNotIn("CLIENT MANAGEMENT CONFIDENTIEL", broad_values)


if __name__ == "__main__":
    unittest.main()
