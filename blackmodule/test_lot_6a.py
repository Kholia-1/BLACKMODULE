"""LOT 6A regression tests for periodic alert quality review."""

import unittest
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import FastAPI, Request
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event, inspect
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.middleware.sessions import SessionMiddleware

from app.database import Base, get_db
from app.models import (
    Alert,
    AlertDecisionHistory,
    AlertQualityReview,
    ApprovalRequest,
    AuditLog,
    User,
)
from app.routers import web
from app.services.alert_decision_service import APPROVED_DECISION, PENDING_DECISION
from app.services.authorization_service import (
    PERMISSION_QUALITY_REVIEW_MANAGE,
    PERMISSION_QUALITY_REVIEW_VIEW,
    ROLE_ADMIN_TECHNIQUE,
    ROLE_ANALYSTE_CONFORMITE,
    ROLE_AUDITEUR,
    ROLE_CONSULTATION,
    ROLE_SUPERVISEUR_CONFORMITE,
    has_permission,
)
from app.services.quality_review_service import (
    REVIEW_NON_COMPLIANT,
    REVIEW_COMPLIANT,
    QualityReviewError,
    build_quality_review_dashboard,
    create_quality_review,
    resolve_quality_review_filters,
)


def _test_app(db) -> FastAPI:
    app = FastAPI()
    app.add_middleware(SessionMiddleware, secret_key="lot-6a-local-test")

    @app.get("/_test/login/{role}")
    def login(request: Request, role: str):
        user = db.query(User).filter(User.role == role).first()
        request.session["user"] = {
            "id": str(user.id),
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role,
        }
        return {"status": "ok"}

    app.dependency_overrides[get_db] = lambda: db
    app.include_router(web.router)
    return app


class Lot6AQualityReviewTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.db = sessionmaker(bind=self.engine, expire_on_commit=False)()
        self.now = datetime(2026, 9, 3, 12, 0)
        self.analyst1 = self.user("analyste-quality-1", ROLE_ANALYSTE_CONFORMITE)
        self.analyst2 = self.user("analyste-quality-2", ROLE_ANALYSTE_CONFORMITE)
        self.admin = self.user("admin-quality", ROLE_ADMIN_TECHNIQUE)
        self.supervisor = self.user("superviseur-quality", ROLE_SUPERVISEUR_CONFORMITE)
        self.auditor = self.user("auditeur-quality", ROLE_AUDITEUR)
        self.reader = self.user("lecteur-quality", ROLE_CONSULTATION)
        self.fp = self.decision("FP-ONU", "ONU", "FAUX_POSITIF", self.analyst1, 12)
        self.confirmed = self.decision("CONF-OFAC", "OFAC_SDN", "CONFIRMEE", self.analyst1, 24)
        self.confirmed_2 = self.decision("CONF-UE", "UE", "CONFIRMEE", self.analyst2, 36)
        self.pending = self.decision(
            "PENDING", "ONU", "FAUX_POSITIF", self.analyst2, 8,
            decision_status=PENDING_DECISION,
        )
        self.old = self.decision("OLD", "ONU", "FAUX_POSITIF", self.analyst2, 40 * 24)
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.clear_compiled_cache()
        self.engine.dispose()

    def user(self, username, role):
        user = User(
            username=username,
            full_name=username.replace("-", " ").title(),
            email=f"{username}@example.test",
            password_hash="local-test",
            role=role,
            statut="ACTIF",
        )
        self.db.add(user)
        self.db.flush()
        return user

    def decision(self, reference, source, status, analyst, hours_ago, *, decision_status=APPROVED_DECISION):
        applied_at = self.now - timedelta(hours=hours_ago)
        alert = Alert(
            client_reference=reference,
            client_nom="CLIENT QUALITE CONFIDENTIEL",
            source_liste=source,
            niveau_alerte="ALERTE_EXACTE",
            statut=status,
            assigned_to_user_id=analyst.id,
            assigned_to=analyst.username,
            created_at=applied_at - timedelta(hours=2),
            treated_at=applied_at,
            treatment_comment="COMMENTAIRE METIER CONFIDENTIEL",
        )
        self.db.add(alert)
        self.db.flush()
        approval = ApprovalRequest(
            operation_type="ALERT_TREATMENT",
            status="VALIDE" if decision_status == APPROVED_DECISION else "EN_ATTENTE_VALIDATION",
            initiator_user_id=str(analyst.id),
            initiated_by=analyst.username,
            reviewer_user_id=str(self.supervisor.id),
            reviewed_by=self.supervisor.username,
            target_entity_type="Alert",
            target_entity_id=str(alert.id),
            initiator_comment="MOTIF APPROBATION CONFIDENTIEL",
            reviewer_comment="COMMENTAIRE SUPERVISEUR CONFIDENTIEL",
            created_at=applied_at - timedelta(hours=1),
            reviewed_at=applied_at if decision_status == APPROVED_DECISION else None,
        )
        self.db.add(approval)
        self.db.flush()
        decision = AlertDecisionHistory(
            alert_id=alert.id,
            approval_request_id=approval.id,
            old_status="EN_COURS",
            requested_status=status,
            decision_status=decision_status,
            initiated_by=analyst.username,
            initiated_at=applied_at - timedelta(hours=1),
            reason="MOTIF DECISION CONFIDENTIEL",
            reviewed_by=self.supervisor.username if decision_status == APPROVED_DECISION else None,
            reviewed_at=applied_at if decision_status == APPROVED_DECISION else None,
            applied_at=applied_at if decision_status == APPROVED_DECISION else None,
        )
        self.db.add(decision)
        self.db.flush()
        return decision

    @staticmethod
    def actor(user):
        return {"id": str(user.id), "username": user.username, "role": user.role}

    def report(self, **kwargs):
        filters = resolve_quality_review_filters(now=self.now, **kwargs)
        return build_quality_review_dashboard(self.db, filters)

    def test_01_review_is_traced_without_changing_business_decision(self):
        decision_before = (self.fp.requested_status, self.fp.decision_status, self.fp.applied_at)
        alert = self.db.get(Alert, self.fp.alert_id)
        alert_before = (alert.statut, alert.treated_at, alert.treated_by)
        reviewed_at = self.now - timedelta(minutes=5)
        review = create_quality_review(
            self.db,
            decision_id=self.fp.id,
            review_status=REVIEW_COMPLIANT,
            quality_comment="Contrôle documentaire conforme.",
            actor=self.actor(self.auditor),
            reviewed_at=reviewed_at,
        )
        self.db.commit()
        self.assertEqual(review.reviewed_by, self.auditor.username)
        self.assertEqual(review.reviewed_at, reviewed_at)
        self.assertEqual(review.alert_id, self.fp.alert_id)
        self.assertEqual(
            (self.fp.requested_status, self.fp.decision_status, self.fp.applied_at),
            decision_before,
        )
        self.assertEqual((alert.statut, alert.treated_at, alert.treated_by), alert_before)

    def test_02_non_compliance_requires_a_separate_bounded_quality_comment(self):
        with self.assertRaises(QualityReviewError):
            create_quality_review(
                self.db, decision_id=self.fp.id, review_status=REVIEW_NON_COMPLIANT,
                quality_comment="  ", actor=self.actor(self.auditor),
            )
        with self.assertRaises(QualityReviewError):
            create_quality_review(
                self.db, decision_id=self.fp.id, review_status=REVIEW_COMPLIANT,
                quality_comment="x" * 2001, actor=self.actor(self.auditor),
            )
        with self.assertRaises(QualityReviewError):
            create_quality_review(
                self.db, decision_id=self.fp.id, review_status=REVIEW_COMPLIANT,
                quality_comment=None,
                actor={"id": "identifiant-invalide", "username": "auditeur", "role": ROLE_AUDITEUR},
            )

    def test_03_history_is_append_only_and_latest_review_drives_status(self):
        first = create_quality_review(
            self.db, decision_id=self.fp.id, review_status=REVIEW_COMPLIANT,
            quality_comment="Première revue.", actor=self.actor(self.auditor),
            reviewed_at=self.now - timedelta(minutes=10),
        )
        second = create_quality_review(
            self.db, decision_id=self.fp.id, review_status=REVIEW_NON_COMPLIANT,
            quality_comment="Anomalie de justification.", actor=self.actor(self.supervisor),
            reviewed_at=self.now - timedelta(minutes=5),
        )
        self.db.commit()
        report = self.report(period="7")
        item = next(row for row in report["items"] if row["decision_id"] == str(self.fp.id))
        self.assertEqual(item["review_status"], REVIEW_NON_COMPLIANT)
        self.assertEqual([row["status"] for row in item["history"]], [REVIEW_NON_COMPLIANT, REVIEW_COMPLIANT])
        self.assertEqual(self.db.query(AlertQualityReview).filter_by(decision_history_id=self.fp.id).count(), 2)
        self.assertNotEqual(first.id, second.id)

    def test_04_kpis_and_filters_are_exact(self):
        create_quality_review(
            self.db, decision_id=self.fp.id, review_status=REVIEW_NON_COMPLIANT,
            quality_comment="Anomalie réelle.", actor=self.actor(self.auditor),
        )
        create_quality_review(
            self.db, decision_id=self.confirmed.id, review_status=REVIEW_COMPLIANT,
            quality_comment=None, actor=self.actor(self.supervisor),
        )
        self.db.commit()
        report = self.report(period="7")
        self.assertEqual(report["kpis"], {
            "total": 3, "pending": 1, "reviewed": 2,
            "compliant": 1, "non_compliant": 1,
            "compliance_rate": 50.0, "non_compliance_rate": 50.0,
        })
        analyst = {row["analyst"]: row for row in report["analyst_kpis"]}
        self.assertEqual(analyst[self.analyst1.username]["total"], 2)
        self.assertEqual(analyst[self.analyst1.username]["compliance_rate"], 50.0)
        self.assertEqual(analyst[self.analyst1.username]["non_compliance_rate"], 50.0)
        self.assertEqual(self.report(period="7", source="ONU")["kpis"]["total"], 1)
        self.assertEqual(self.report(period="7", analyst=self.analyst2.username)["kpis"]["total"], 1)
        self.assertEqual(self.report(period="7", decision="FAUX_POSITIF")["kpis"]["total"], 1)
        self.assertEqual(self.report(period="7", review_status="NON_CONFORME")["kpis"]["total"], 1)
        self.assertEqual(self.report(period="7", review_status="A_REVOIR")["kpis"]["total"], 1)
        self.assertEqual(self.report(period="30")["kpis"]["total"], 3)

    def test_05_pending_and_non_reviewable_decisions_are_excluded(self):
        closed = self.decision("CLOSED", "ONU", "CLOTUREE", self.analyst1, 3)
        self.db.commit()
        report = self.report(period="7")
        ids = {item["decision_id"] for item in report["items"]}
        self.assertNotIn(str(self.pending.id), ids)
        self.assertNotIn(str(closed.id), ids)
        with self.assertRaises(QualityReviewError):
            create_quality_review(
                self.db, decision_id=self.pending.id, review_status=REVIEW_COMPLIANT,
                quality_comment=None, actor=self.actor(self.auditor),
            )

    def test_06_rbac_allows_existing_admin_and_supervisor_but_not_operator_or_reader(self):
        self.assertTrue(has_permission(self.actor(self.admin), PERMISSION_QUALITY_REVIEW_VIEW))
        self.assertTrue(has_permission(self.actor(self.admin), PERMISSION_QUALITY_REVIEW_MANAGE))
        self.assertTrue(has_permission(self.actor(self.auditor), PERMISSION_QUALITY_REVIEW_VIEW))
        self.assertTrue(has_permission(self.actor(self.auditor), PERMISSION_QUALITY_REVIEW_MANAGE))
        self.assertTrue(has_permission(self.actor(self.supervisor), PERMISSION_QUALITY_REVIEW_VIEW))
        self.assertTrue(has_permission(self.actor(self.supervisor), PERMISSION_QUALITY_REVIEW_MANAGE))
        self.assertFalse(has_permission(self.actor(self.analyst1), PERMISSION_QUALITY_REVIEW_VIEW))
        self.assertFalse(has_permission(self.actor(self.reader), PERMISSION_QUALITY_REVIEW_MANAGE))
        with TestClient(_test_app(self.db)) as client:
            for role in (ROLE_ADMIN_TECHNIQUE, ROLE_SUPERVISEUR_CONFORMITE, ROLE_AUDITEUR):
                client.get(f"/_test/login/{role}")
                response = client.get(
                    "/web/quality-review?period=custom&date_from=2026-09-01&date_to=2026-09-03"
                )
                self.assertEqual(response.status_code, 200)
                self.assertIn("Revue qualité périodique", response.text)
                self.assertIn("Revue qualité", response.text)
                self.assertIn("Taux de conformité", response.text)
                self.assertIn("Taux de non-conformité", response.text)
                self.assertEqual(client.post(
                    f"/web/quality-review/{self.fp.id}",
                    data={"review_status": REVIEW_COMPLIANT}, follow_redirects=False,
                ).status_code, 303)
            for role in (ROLE_ANALYSTE_CONFORMITE, ROLE_CONSULTATION):
                client.get(f"/_test/login/{role}")
                self.assertEqual(client.get("/web/quality-review").status_code, 403)
                self.assertEqual(client.get("/web/quality-review/export.csv").status_code, 403)
                self.assertEqual(client.post(
                    f"/web/quality-review/{self.fp.id}",
                    data={"review_status": REVIEW_COMPLIANT},
                ).status_code, 403)

    def test_07_service_prevents_analyst_self_review(self):
        with self.assertRaises(PermissionError):
            create_quality_review(
                self.db, decision_id=self.fp.id, review_status=REVIEW_COMPLIANT,
                quality_comment=None, actor=self.actor(self.analyst1),
            )

    def test_08_web_review_audit_is_generic_and_business_data_unchanged(self):
        status_before = self.db.get(Alert, self.fp.alert_id).statut
        with TestClient(_test_app(self.db)) as client:
            client.get(f"/_test/login/{ROLE_AUDITEUR}")
            response = client.post(
                f"/web/quality-review/{self.fp.id}",
                data={
                    "review_status": REVIEW_NON_COMPLIANT,
                    "quality_comment": "COMMENTAIRE QUALITE CONFIDENTIEL",
                },
                follow_redirects=False,
            )
        self.assertEqual(response.status_code, 303)
        audit = self.db.query(AuditLog).filter_by(action="QUALITY_REVIEW_RECORDED").one()
        self.assertEqual(audit.description, "Revue qualité enregistrée.")
        self.assertNotIn("CONFIDENTIEL", audit.description)
        self.assertEqual(self.db.get(Alert, self.fp.alert_id).statut, status_before)
        self.assertEqual(self.db.get(AlertDecisionHistory, self.fp.id).decision_status, APPROVED_DECISION)

    def test_09_csv_and_xlsx_exports_respect_filters_and_hide_sensitive_text(self):
        create_quality_review(
            self.db, decision_id=self.fp.id, review_status=REVIEW_NON_COMPLIANT,
            quality_comment="COMMENTAIRE QUALITE CONFIDENTIEL",
            actor=self.actor(self.auditor),
        )
        self.db.commit()
        with TestClient(_test_app(self.db)) as client:
            client.get(f"/_test/login/{ROLE_AUDITEUR}")
            csv_response = client.get(
                "/web/quality-review/export.csv?period=custom&date_from=2026-09-01&date_to=2026-09-03&source=ONU"
            )
            xlsx_response = client.get(
                "/web/quality-review/export.xlsx?period=custom&date_from=2026-09-01&date_to=2026-09-03&source=ONU"
            )
        self.assertEqual(csv_response.status_code, 200)
        self.assertIn("text/csv", csv_response.headers["content-type"])
        self.assertIn("ONU", csv_response.text)
        self.assertNotIn("OFAC_SDN", csv_response.text)
        self.assertNotIn("CLIENT QUALITE CONFIDENTIEL", csv_response.text)
        self.assertNotIn("COMMENTAIRE QUALITE CONFIDENTIEL", csv_response.text)
        self.assertNotIn("MOTIF DECISION CONFIDENTIEL", csv_response.text)
        workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
        worksheet_text = "\n".join(
            str(value) for row in workbook.active.iter_rows(values_only=True)
            for value in row if value is not None
        )
        self.assertIn("ONU", worksheet_text)
        self.assertNotIn("OFAC_SDN", worksheet_text)
        self.assertNotIn("CONFIDENTIEL", worksheet_text)
        self.assertEqual(self.db.query(AuditLog).filter(AuditLog.action.like("EXPORT_QUALITY_REVIEW_%")).count(), 2)

    def test_10_dashboard_has_constant_queries_and_never_loads_alert_entities(self):
        query_count = 0
        loaded_alerts = 0

        def count_query(*_args, **_kwargs):
            nonlocal query_count
            query_count += 1

        def alert_loaded(*_args, **_kwargs):
            nonlocal loaded_alerts
            loaded_alerts += 1

        self.db.expunge_all()
        event.listen(self.engine, "before_cursor_execute", count_query)
        event.listen(Alert, "load", alert_loaded)
        try:
            report = self.report(period="7")
        finally:
            event.remove(self.engine, "before_cursor_execute", count_query)
            event.remove(Alert, "load", alert_loaded)
        self.assertEqual(report["kpis"]["total"], 3)
        self.assertEqual(query_count, 5)
        self.assertEqual(loaded_alerts, 0)

    def test_11_schema_creation_is_idempotent_and_indexes_exist(self):
        Base.metadata.create_all(self.engine)
        Base.metadata.create_all(self.engine)
        inspector = inspect(self.engine)
        self.assertIn("alert_quality_reviews", inspector.get_table_names())
        columns = {column["name"]: column for column in inspector.get_columns("alert_quality_reviews")}
        self.assertEqual(
            set(columns),
            {
                "id", "alert_id", "decision_history_id", "review_status",
                "quality_comment", "reviewed_by_user_id", "reviewed_by", "reviewed_at",
            },
        )
        self.assertFalse(columns["review_status"]["nullable"])
        indexes = {index["name"] for index in inspector.get_indexes("alert_quality_reviews")}
        self.assertIn("ix_alert_quality_decision_reviewed", indexes)
        self.assertIn("ix_alert_quality_reviewer_reviewed", indexes)

    def test_12_quality_review_history_rejects_updates_and_deletes(self):
        review = create_quality_review(
            self.db, decision_id=self.fp.id, review_status=REVIEW_COMPLIANT,
            quality_comment="Revue initiale.", actor=self.actor(self.auditor),
        )
        self.db.commit()
        review.review_status = REVIEW_NON_COMPLIANT
        with self.assertRaisesRegex(ValueError, "immuables"):
            self.db.flush()
        self.db.rollback()
        persisted = self.db.get(AlertQualityReview, review.id)
        self.assertEqual(persisted.review_status, REVIEW_COMPLIANT)
        self.db.delete(persisted)
        with self.assertRaisesRegex(ValueError, "immuables"):
            self.db.flush()
        self.db.rollback()
        self.assertEqual(self.db.query(AlertQualityReview).filter_by(id=review.id).count(), 1)


if __name__ == "__main__":
    unittest.main()
